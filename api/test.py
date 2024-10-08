import unittest 
import io
#from dash_page import download_csv
from openpyxl import load_workbook
import pandas as pd


class TestDownloadxlsx(unittest.TestCase):

    def setUp(self):
        self.expected_file_path = './api/tests/CSL_3_D.xlsx'
        self.test_file_path = './api/tests/CSL_3_D_t.xlsx'

    def find_start_row(self, sheet):
        for index in range(1, sheet.max_row + 1):
            if sheet[index][0].value in ["Stage no.", "time start of stage"]:
                return index

    def test_download_xlsx_function(self):

        # Load Excel files 
        expected_file = load_workbook(self.expected_file_path)["03 - Shearing"]
        test_file = load_workbook(self.test_file_path)["Shearing"]

        # Find start row
        expected_start_row = self.find_start_row(expected_file)
        test_start_row = self.find_start_row(test_file)
            
        # Create dataframes of Excel sheets for comparison 
        expected_content = pd.read_excel(self.expected_file_path, sheet_name = "03 - Shearing", skiprows = expected_start_row-1).drop([0])
        test_content = pd.read_excel(self.test_file_path, sheet_name = "Shearing", skiprows = test_start_row-1)
        
        # Compare values in p' column 
        self.assertAlmostEqual(expected_content["p'"].iloc[-1], test_content["p'"].iloc[-1], places=10)
        print("p' column matches")

        # Compare values in axial strain column 
        self.assertAlmostEqual(expected_content["Axial strain"].iloc[-1], test_content["axial strain"].iloc[-1], places=10)
        print("Axial strain column matches")

        # Compare values in e column 
        self.assertAlmostEqual(expected_content["Void ratio"].iloc[-1], test_content["void ratio"].iloc[-1], places=10)
        print("Void ratio column matches")

if __name__ == '__main__':
    unittest.main()