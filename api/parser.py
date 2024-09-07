import pandas as pd
import openpyxl


"""
Wanted columns:
    - Time start of stage
    - Axial strain
    - Volumetric strain
    - Excess PWP
    - p'
    - Deviator stress
    - Void ratio
    - Shear induced PWP
"""

# Cutoff row for search for table start cell
CUTOFF = 50

columns = {
    "time start of stage",
    "axial strain",
    "volumetric strain",
    "excess pwp",
    "p'",
    "deviator stress",
    "void ratio",
    "shear induced pwp"
}

specs = {
    "drainage",
    "shearing",
    "anisotropy",
    "consolidation",
    "availability",
    "density",
    "plasticity",
    "psd"
}



## TODO: check for errors, some labelled ERROR_CHECK

def load_sheet(sheet_path):
    
    # Uses openpyxl to open and read a workbook using data_only mode to read results of eqs
    wb = openpyxl.load_workbook(filename=sheet_path, data_only=True, read_only=True)
    
    # Loads specific sheet object out of workbook
    sheet = wb[wb.sheetnames[3]]
    
    return sheet


def ingest_specs(sheet) -> dict:

    # counts number of hits
    keys_populated = set()
    spec_dict = dict()

    for _, *values in sheet.iter_rows(max_row=CUTOFF):
        for i, v in enumerate(values):

            cell_value = str(v.value).lower().strip()

            if cell_value in specs:

                if cell_value in keys_populated:
                    return "error"      # ERROR_CHECK
                
                spec_dict[cell_value] = str(values[i+1].value).lower().strip()
                keys_populated.add(cell_value)

    return spec_dict


def ingest_table(sheet) -> pd.DataFrame:
    
    # Convert sheet into DataFrame object
    df = pd.DataFrame(sheet.values)

    # We first need to detect the start of the table within the sheet using "Stage no."
    for i in range(CUTOFF):
        if type(df.iloc[0][0]) is str and df.iloc[0][0].strip().lower() == "stage no.":
            df = df.reset_index(drop=True)
            break
        else:
            df = df.drop([i])

    # Identify the columns that need to be dropped from the DataFrame
    to_drop = []
    for i, name in enumerate(df.iloc[0]):
        if type(name) != str or name.lower().strip() not in columns:
            to_drop.append(i)

    df = df.drop(df.columns[to_drop], axis=1)
    
    # Set column names
    df.columns = df.iloc[0]
    # Strip whitespace from column names
    df = df.rename(columns = lambda x: x.strip())
    
    # Drop column names from data body after assignment
    df = df.drop([0,1])
    df = df.reset_index(drop=True)

    # Find end of table and remove None values if excess cells occur in Excel file
    end_table_index = 0
    for i, value in enumerate(df[df.columns[0]]):
        if value == None:
            end_table_index = i
            break

    df = df[:end_table_index]

    return df


def parse_workbook(path):
    
    sheet = load_sheet(path)

    specs = ingest_specs(sheet)
    df = ingest_table(sheet)

    return specs, df



## TESTING
"""
specs, df = parse_workbook("./CSL_1_U.xlsx")

print(df)
"""
