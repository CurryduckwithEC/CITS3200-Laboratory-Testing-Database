import pandas as pd
import openpyxl


"""
Wanted columns:
    - Time start of stage
    - Axial strain
    - Volumetric strain
    - Excess PWP
    - p'
    - Deviator stress (q)
    - Void ratio (w)
    - Shear induced PWP
"""

# Cutoff row for search for table start cell
CUTOFF = 50

columns = {
    "time start of stage",
    "axial strain",
    "volumetric strain",
    "excess pwp",
    "p'",
    "deviator stress",
    "void ratio",
    "shear induced pwp"
}

specs = {
    "drainage",
    "shearing",
    "anisotropy",
    "consolidation",
    "availability",
    "density",
    "plasticity",
    "psd"
}



## TODO: check for errors, some labelled ERROR_CHECK

def load_sheet(sheet_path):
    
    # Uses openpyxl to open and read a workbook using data_only mode to read results of eqs
    wb = openpyxl.load_workbook(filename=sheet_path, data_only=True, read_only=True)
    
    # Loads specific sheet object out of workbook
    sheet = wb[wb.sheetnames[3]]
    
    return sheet


def ingest_specs(sheet) -> dict:

    # counts number of hits
    keys_populated = set()
    spec_dict = dict()

    for _, *values in sheet.iter_rows(max_row=CUTOFF):
        for i, v in enumerate(values):

            cell_value = str(v.value).lower().strip()

            if cell_value in specs:

                if cell_value in keys_populated:
                    return "error"      # ERROR_CHECK
                
                spec_dict[cell_value] = str(values[i+1].value).lower().strip()
                keys_populated.add(cell_value)

    return spec_dict


def ingest_table(sheet) -> pd.DataFrame:
    
    # Convert sheet into DataFrame object
    df = pd.DataFrame(sheet.values)

    # We first need to detect the start of the table within the sheet using "Stage no."
    for i in range(CUTOFF):
        if type(df.iloc[0][0]) is str and df.iloc[0][0].strip().lower() == "stage no.":
            df = df.reset_index(drop=True)
            break
        else:
            df = df.drop([i])

    # Identify the columns that need to be dropped from the DataFrame
    to_drop = []
    for i, name in enumerate(df.iloc[0]):
        if type(name) != str or name.lower().strip() not in columns:
            to_drop.append(i)

    df = df.drop(df.columns[to_drop], axis=1)
    
    # Set column names
    df.columns = df.iloc[0]
    
    # Drop column names from data body after assignment
    df = df.drop([0,1])
    df = df.reset_index(drop=True)

    # Find end of table and remove None values if excess cells occur in Excel file
    end_table_index = 0
    for i, value in enumerate(df[df.columns[0]]):
        if value == None:
            end_table_index = i
            break

    df = df[:end_table_index]

    return df


def parse_workbook(path):
    
    sheet = load_sheet(path)

    specs = ingest_specs(sheet)
    df = ingest_table(sheet)

    return specs, df

### VISUALIATIONS

specs, df = parse_workbook("./CSL_1_U.xlsx")
specs2, df2 = parse_workbook("./CSL_2_U.xlsx")
specs3, df3 = parse_workbook("./CSL_3_D.xlsx")

# Add Test column to differenciate different tests
df['Test'] = 'Test1'
df2['Test'] = 'Test2'
df3['Test'] = 'Test3'

df_combined = pd.concat([df, df2, df3])

df_combined["Stress ratio"] = df_combined["Deviator stress"]/df_combined["p'"]

#print(df_combined)

import plotly.express as px 

### Deviator Stress & Mean effective stress (p') VS Axial Strain 
axial_deviator_fig = px.line(
    df_combined, x="Axial strain", y=["Deviator stress", "p'"], labels={'x': 'Axial strain', 'value':"Deviator Stress & Mean Effective Stress, p'"}, color="Test", title="Deviator and Mean Effective Stress (kPa) vs. Axial Strain (%)")
axial_deviator_fig.write_html('./plots/axial_deviator_fig.html')

### Shear induced PWP VS Axial Strain
axial_pwp = px.line(
    df_combined, x="Axial strain", y="Shear induced PWP",color="Test", title="Shear Induced Pore Pressure (kPa) vs. Axial Strain (%)")
axial_pwp.write_html('./plots/axial_pwp.html')

### Deviator Stress (q) VS Mean effective stress (p')
q_p = px.line(
    df_combined, x="p'", y="Deviator stress",color="Test", title="Deviator Stress, q (kPa) vs. Mean Effective Stress, p' (kPa)")
q_p.write_html('./plots/q_p.html')

### Volumetric Strain VS Axial Strain
axial_vol = px.line(
    df_combined, x="Axial strain", y="Volumetric strain",color="Test", title="Volumetric Stress (%) vs. Axial Strain (%)")
axial_vol.write_html('./plots/axial_vol.html')

### Stress ratio (q/p') VS Axial Strain
axial_stressratio = px.line(
    df_combined, x="Axial strain", y="Stress ratio",color="Test", title="Stress Ratio, q/p' vs. Axial Strain (%)")
axial_stressratio.write_html('./plots/axial_stressratio.html')

### Deviator Stress (q) VS Mean Effective Stress
#axial_stressratio = px.line(df_combined, x="Axial strain", y="Stress ratio",color="Test")
#axial_stressratio.write_html('./plots/time_pwp.html')


### Axial Strain VS Time 
# Some issues with the column "Time start of stage", error says that is expected 'Time start of stage ' with space at the back
#time_axial_vol = px.line(df_combined, x='Time start of stage ', y=["Axial strain", "Volumetric strain"],color="Test")
#time_axial_vol.write_html('./plots/time_axial_vol.html')
