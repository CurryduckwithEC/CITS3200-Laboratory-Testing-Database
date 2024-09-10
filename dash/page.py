# See official docs at https://dash.plotly.com
# pip install dash pandas

from dash import Dash, dcc, html, Input, Output, ctx, callback
import plotly.express as px
import pandas as pd
import sys
import openpyxl
import dash_bootstrap_components as dbc
import numpy as np 

### Read Data 

"""
Wanted columns:
    - Time start of stage
    - Axial strain
    - Volumetric strain
    - Excess PWP
    - p'
    - Deviator stress (q)
    - Void ratio (w)
    - Shear induced PWP
"""

# Cutoff row for search for table start cell
CUTOFF = 50

columns = {
    "time start of stage",
    "axial strain",
    "volumetric strain",
    "excess pwp",
    "p'",
    "deviator stress",
    "void ratio",
    "shear induced pwp"
}

specs = {
    "drainage",
    "shearing",
    "anisotropy",
    "consolidation",
    "availability",
    "density",
    "plasticity",
    "psd"
}

def load_sheet(sheet_path):
    
    # Uses openpyxl to open and read a workbook using data_only mode to read results of eqs
    wb = openpyxl.load_workbook(filename=sheet_path, data_only=True, read_only=True)
    
    # Loads specific sheet object out of workbook
    sheet = wb[wb.sheetnames[3]]
    
    return sheet


def ingest_specs(sheet) -> dict:

    # counts number of hits
    keys_populated = set()
    spec_dict = dict()

    for _, *values in sheet.iter_rows(max_row=CUTOFF):
        for i, v in enumerate(values):

            cell_value = str(v.value).lower().strip()

            if cell_value in specs:

                if cell_value in keys_populated:
                    return "error"      # ERROR_CHECK
                
                spec_dict[cell_value] = str(values[i+1].value).lower().strip()
                keys_populated.add(cell_value)

    return spec_dict


def ingest_table(sheet) -> pd.DataFrame:
    
    # Convert sheet into DataFrame object
    df = pd.DataFrame(sheet.values)

    # We first need to detect the start of the table within the sheet using "Stage no."
    for i in range(CUTOFF):
        if type(df.iloc[0][0]) is str and df.iloc[0][0].strip().lower() == "stage no.":
            df = df.reset_index(drop=True)
            break
        else:
            df = df.drop([i])

    # Identify the columns that need to be dropped from the DataFrame
    to_drop = []
    for i, name in enumerate(df.iloc[0]):
        if type(name) != str or name.lower().strip() not in columns:
            to_drop.append(i)

    df = df.drop(df.columns[to_drop], axis=1)
    
    # Set column names
    df.columns = df.iloc[0]
    
    # Drop column names from data body after assignment
    df = df.drop([0,1])
    df = df.reset_index(drop=True)

    # Find end of table and remove None values if excess cells occur in Excel file
    end_table_index = 0
    for i, value in enumerate(df[df.columns[0]]):
        if value == None:
            end_table_index = i
            break

    df = df[:end_table_index]

    return df


def parse_workbook(path):
    
    sheet = load_sheet(path)

    specs = ingest_specs(sheet)
    df = ingest_table(sheet)

    return specs, df

specs, df = parse_workbook("./data/CSL_1_U.xlsx")
specs2, df2 = parse_workbook("./data/CSL_2_U.xlsx")
specs3, df3 = parse_workbook("./data/CSL_3_D.xlsx")

print(specs2)

# Add Test column to differenciate tests
df['Test'] = 'Test1'
df2['Test'] = 'Test2'
df3['Test'] = 'Test3'

df_combined = pd.concat([df, df2, df3])

# Stress ratio column 
df_combined["Stress ratio"] = df_combined["Deviator stress"]/df_combined["p'"]

# log(p') column, forcing a NaN if p' is invalid 
df_combined["log(p')"] = np.log(pd.to_numeric(df_combined["p'"], errors="coerce"))

### Create Visualisations

app = Dash(__name__)

app.layout = html.Div(
    id = "app-container",
    style = {
        "display": "flex",
        "margin": "5px"},
    children = [   
        html.Div(
            id = "filters",
            style = {
                "backgroundColor": "lightGrey",
                "padding": "5px",
                "float": "left"
            },
            children = [
                html.Div(
                    dbc.Accordion(
                        [
                            dbc.AccordionItem(
                                [
                                    # Have to go back and put in actual values used in DB
                                    # Format Label:Value
                                    html.H2("Drainage"),
                                    dcc.Checklist(
                                        options={
                                            "Drained":"Drained", 
                                            "Undrained":"Undrained"},
                                        value=["Drained", "Undrained"],
                                        id="drainage_checklist",
                                        inline=True
                                    ),
                                    html.H2("Shearing"),
                                    dcc.Checklist(
                                        options={
                                            "Compression":"Compression", 
                                            "Extension":"Extension"},
                                        value=["Compression", "Extension"],
                                        id="shearing_checklist",
                                        inline=True
                                    ),
                                    html.H2("Anisotropy"),
                                    dcc.Checklist(
                                        options={
                                            "Isotropic":"Isotropic", 
                                            "Anisotropic":"Anisotropic"},
                                        value=["Isotropic", "Anisotropic"],
                                        id="anisotropy_checklist",
                                        inline=True
                                    ),
                                    dcc.RangeSlider(
                                        0.3,
                                        1,
                                        step=None,
                                        value=[0.3,1.0], 
                                        id="anisotropy_slider",
                                        tooltip={"placement": "bottom", "always_visible": True}
                                    ),
                                    "Min Value:",
                                    dcc.Input(id="anisotropy_min_value", type="number", min=0, max=1.0, value=1.0, style={'width': '80px'}),
                                    "Max Value:",
                                    dcc.Input(id="anisotropy_max_value", type="number", min=0, max=1.0, value=0, style={'width': '80px'}),
                                    
                                    html.H2("Consolidation"),
                                    html.P(id="consolidation_value"),
                                    dcc.RangeSlider(
                                        10,
                                        1500,
                                        step=1,
                                        marks={
                                            10: "10",
                                            300: "300",
                                            600: "600",
                                            900: "900",
                                            1200: "1200",
                                            1500: "1500"
                                        },
                                        value=[10,1500], 
                                        id="consolidation_slider",
                                        tooltip={"placement": "bottom", "always_visible": True}
                                    ),
                                    "Min Value:",
                                    dcc.Input(id="consolidation_min_value", type="number", min=0, max=1500, value=1500, style={'width': '80px'}),
                                    "Max Value:",
                                    dcc.Input(id="consolidation_max_value", type="number", min=0, max=1500, value=0, style={'width': '80px'}),

                                    html.H2("Availability"),
                                    dcc.Checklist(
                                        options={
                                            "Public":"Public", 
                                            "Confidential":"Confidential"},
                                        value=["Public", "Confidential"],
                                        id="availability_checklist",
                                        inline=True
                                    )
                                ],
                                title="Test",
                            ),
                            dbc.AccordionItem(
                                [
                                    html.H2("Density"),
                                    dcc.Checklist(
                                        options={
                                            "Loose":"Loose",
                                            "Dense":"Dense"},
                                        value=["Loose", "Dense"],
                                        id="density_checklist",
                                        inline=True
                                    ),
                                    html.H2("Plasticity"),
                                    dcc.Checklist(
                                        options={
                                            "Plastic":"Plastic",
                                            "Non-plastic":"Nonplastic",
                                            "Unknown":"Unknown"},
                                        value=["Plastic", "Non-plastic","Unknown"],
                                        id="plasticity_checklist",
                                        inline=True
                                    ),
                                    html.H2("PSD"),
                                    dcc.Checklist(
                                        options={
                                            "Clay":"Clay",
                                            "Sand":"Sand",
                                            "Silt":"Silt"},
                                        value=["Clay", "Sand","Silt"],
                                        id="psd_checklist",
                                        inline=True
                                    ),
                                ],
                                title="Sample",
                            ),
                            dbc.AccordionItem(
                                [
                                    html.H2("Axial Strain Filter"),
                                    html.P(id="axial_value"),
                                    dcc.RangeSlider(
                                        0,
                                        0.5,
                                        step=0.05,
                                        marks={
                                            0.0: "0.0",
                                            0.1: "0.1",
                                            0.2: "0.2",
                                            0.3: "0.3",
                                            0.4: "0.4",
                                            0.5: "0.5"
                                        },
                                        value=[0,0.5], 
                                        id='axial_slider'
                                    ),
                                    "Min Value:",
                                    dcc.Input(id="axial_min_value", type="number", min=0, max=0.5, value=0.5, style={'width': '80px'}),
                                    "Max Value:",
                                    dcc.Input(id="axial_max_value", type="number", min=0, max=0.5, value=0, style={'width': '80px'}),
                                    
                                    html.H2("p' Filter"),
                                    html.P(id="p_value"),
                                    dcc.RangeSlider(
                                        0,
                                        500,
                                        step=None,
                                        value=[0,500], 
                                        id='p_slider'
                                    ), 
                                    html.H2("Induced PWP Filter"),
                                    html.P(id="pwp_value"),
                                    dcc.RangeSlider(
                                        0,
                                        500,
                                        step=None,
                                        value=[0,500], 
                                        id='pwp_slider'
                                    ),
                                    html.H2("Deviator stress (q) Filter"),
                                    html.P(id="q_value"),
                                    dcc.RangeSlider(
                                        0,
                                        500,
                                        step=None,
                                        value=[0,500], 
                                        id='q_slider'
                                    ),

                                    html.H2("e Filter"),
                                    html.P(id="e_value"),
                                    dcc.RangeSlider(
                                        0,
                                        1,
                                        step=0.01,
                                        marks={
                                            0.0: "0.0",
                                            0.2: "0.2",
                                            0.4: "0.4",
                                            0.6: "0.6",
                                            0.8: "0.8",
                                            1.0: "1.0"
                                        },
                                        value=[0,1], 
                                        id='e_slider'
                                    ),
                                    ],
                                title="Variables",
                            ),
                        ],
                        always_open=True, flush=True,
                    )
                ),
                
            ]
        ), 
        html.Div(
            id = "dashboard",
            children = [
                dcc.Graph(id="axial_deviator_fig"),
                dcc.Graph(id="axial_pwp_fig"), 
                dcc.Graph(id="q_p_fig"),
                dcc.Graph(id="axial_vol_fig"), 
                dcc.Graph(id="e_logp_fig")
            ]
        )
    ]
)

@app.callback(
    [
    Output("consolidation_min_value","value"),
    Output("consolidation_max_value","value"),
    Output("consolidation_slider","value")    
    ],    
    [
     Input("consolidation_min_value","value"),
     Input("consolidation_max_value","value"),
     Input("consolidation_slider","value")
    ]   
)
def sync_consol_slider(start, end, slider):
    trigger_id = ctx.triggered_id

    consolidation_min_value = start if trigger_id == "consolidation_min_value" else slider[0]
    consolidation_max_value = end if trigger_id == "consolidation_max_value" else slider[1]
    consolidation_slider_value = slider if trigger_id == "consolidation_slider" else [consolidation_min_value, consolidation_max_value]

    return consolidation_min_value, consolidation_max_value, consolidation_slider_value

@app.callback(
    [
    Output("anisotropy_min_value","value"),
    Output("anisotropy_max_value","value"),
    Output("anisotropy_slider","value")    
    ],    
    [
     Input("anisotropy_min_value","value"),
     Input("anisotropy_max_value","value"),
     Input("anisotropy_slider","value")
    ]   
)
def sync_aniso_slider(start, end, slider):
    trigger_id = ctx.triggered_id

    anisotropy_min_value = start if trigger_id == "anisotropy_min_value" else slider[0]
    anisotropy_max_value = end if trigger_id == "anisotropy_max_value" else slider[1]
    anisotropy_slider_value = slider if trigger_id == "anisotropy_slider" else [anisotropy_min_value, anisotropy_max_value]

    return anisotropy_min_value, anisotropy_max_value, anisotropy_slider_value

@app.callback(
    [
    Output("axial_min_value","value"),
    Output("axial_max_value","value"),
    Output("axial_slider","value")    
    ],    
    [
     Input("axial_min_value","value"),
     Input("axial_max_value","value"),
     Input("axial_slider","value")
    ]   
)
def sync_axial_slider(start, end, slider):
    trigger_id = ctx.triggered_id

    axial_min_value = start if trigger_id == "axial_min_value" else slider[0]
    axial_max_value = end if trigger_id == "axial_max_value" else slider[1]
    axial_slider_value = slider if trigger_id == "axial_slider" else [axial_min_value, axial_max_value]

    return axial_min_value, axial_max_value, axial_slider_value

@app.callback(
    [
    Output("axial_deviator_fig", "figure"),
    Output("axial_pwp_fig", "figure"), 
    Output("q_p_fig", "figure"),
    Output("axial_vol_fig", "figure"), 
    Output("e_logp_fig", "figure")
    ],
    [Input("axial_slider", "value"), 
     Input("p_slider", "value"), 
     Input("pwp_slider", "value"), 
     Input("q_slider", "value"),
     Input("e_slider", "value")
     ], 
     )
def update_figure(selected_axial, selected_p, selected_pwp, selected_q, selected_e):
    filtered_df = df_combined[
        (df_combined["Axial strain"]>=selected_axial[0]) & (df_combined["Axial strain"]<=selected_axial[1])
        & (df_combined["p'"]>=selected_p[0]) & (df_combined["p'"]<=selected_p[1])
        & (df_combined["Shear induced PWP"]>=selected_pwp[0]) & (df_combined["Shear induced PWP"]<=selected_pwp[1])
        & (df_combined["Deviator stress"]>=selected_q[0]) & (df_combined["Deviator stress"]<=selected_q[1])
        & (df_combined["Void ratio"]>=selected_e[0]) & (df_combined["Void ratio"]<=selected_e[1])]

    # Deviator Stress (q) & Mean effective stress (p') VS Axial Strain 
    axial_deviator_fig = px.line(
        filtered_df, 
        x="Axial strain", 
        y=["Deviator stress", "p'"], 
        labels={'x': 'Axial strain', 'value':"Deviator Stress & Mean Effective Stress, p'"}, 
        color="Test", 
        title="Deviator, q and Mean Effective Stress (kPa), p' vs. Axial Strain (%)")

    # Shear induced PWP VS Axial Strain
    axial_pwp_fig = px.line(
        filtered_df, 
        x="Axial strain", 
        y="Shear induced PWP",
        color="Test", 
        title="Shear Induced Pore Pressure (kPa) vs. Axial Strain (%)")
    
    # Deviator Stress (q) VS Mean effective stress (p')
    q_p_fig = px.line(
        filtered_df, 
        x="p'", 
        y="Deviator stress",
        color="Test", 
        title="Deviator Stress, q (kPa) vs. Mean Effective Stress, p' (kPa)")
    
    ### Volumetric Strain VS Axial Strain
    axial_vol_fig = px.line(
        filtered_df, 
        x="Axial strain", 
        y="Volumetric strain",
        color="Test", 
        title="Volumetric Stress (%) vs. Axial Strain (%)")
    
    ### e VS log(p')
    e_logp_fig = px.line(
        filtered_df, 
        x="log(p')", 
        y="Void ratio",
        color="Test", 
        title="e vs. log(p')")
    
    return axial_deviator_fig, axial_pwp_fig, q_p_fig, axial_vol_fig, e_logp_fig

@app.callback(
    [Output("axial_value", "children"),
     Output("p_value", "children"),
     Output("pwp_value", "children"), 
     Output("q_value", "children"),
     Output("e_value", "children")],
    [Input("axial_slider", "value"), 
     Input("p_slider", "value"), 
     Input("pwp_slider", "value"), 
     Input("q_slider", "value"),
     Input("e_slider", "value")]
     )
def update_filters(selected_axial, selected_p, selected_pwp, selected_q, selected_e): 
    return f'Selected range: {selected_axial[0]} to {selected_axial[1]}', f'Selected range: {selected_p[0]} to {selected_p[1]}', f'Selected range: {selected_pwp[0]} to {selected_pwp[1]}', f'Selected range: {selected_q[0]} to {selected_q[1]}', f'Selected range: {selected_e[0]} to {selected_e[1]}'

port = "18019"


if len(sys.argv) > 1:
    port = sys.argv[1]

app.run_server(port=port, debug=True)
